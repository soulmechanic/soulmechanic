from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Protection , Font, PatternFill
from win32com.client.gencache import EnsureDispatch
from openpyxl.worksheet.table import Table
from openpyxl.utils import get_column_letter
# from openpyxl.workbook import workbookPassword
import pandas as pd
import string
import time
import os

def add_data_validation_to_column(LOB, attribute_df, mapping_df,DropDownMapping_df, password):   
    wb = Workbook()     
    ws = wb.create_sheet(LOB, 0)    
    hidden_ws = wb['Sheet']    
    hidden_ws.title = 'MappingSheet'    
    for r in dataframe_to_rows(attribute_df, index=False, header=True):
        ws.append(r)
    for r in dataframe_to_rows(DropDownMapping_df, index=False, header=True):
        hidden_ws.append(r)        
    wb.create_named_range('PROCESSING_OTHERS', hidden_ws , '$B$2:$B$6')
    wb.create_named_range('PROCESSING', hidden_ws , '$B$7:$B$10')
    wb.create_named_range('NON_PROCESSING', hidden_ws , '$B$11:$B$20')
    wb.create_named_range('PROCESSING_OTHERS_CATEGORY', hidden_ws , '$C$2:$C$6')
    f_mapping_df = mapping_df[mapping_df['DropDownOptions'].notnull()]
    for index, rows in f_mapping_df.iterrows():        
        if not rows['DropDownOptions']=='Named Range':
            listToStr = '\"' + str(rows['DropDownOptions']) + '\"'
            # Create a data validation object
            dv = DataValidation(type="list", formula1=listToStr, error='Your entry is not valid',
                                errorTitle='Invalid Entry', prompt='Please select from the list', 
                                promptTitle='Select Option')            
            # Add the data validation to the worksheet
            ws.add_data_validation(dv)
            # Apply the data validation to the column
            for row in ws[rows['Position']]:
                dv.add(row)                    
        elif (rows['DropDownOptions']=='Named Range') & (rows['Columns'] == 'Role'):            
            # Create a data validation object
            dv = DataValidation(type="list", formula1='INDIRECT(SUBSTITUTE(P1," ","_"))', 
                                allow_blank=True,error='Your entry is not valid',errorTitle='Invalid Entry', 
                                prompt='Please select from the list', promptTitle='Select Option')                             
            # Add the data validation to the worksheet
            ws.add_data_validation(dv)
            # Apply the data validation to the column
            for row in ws[rows['Position']]:
                dv.add(row)                
        elif (rows['DropDownOptions']=='Named Range') & (rows['Columns'] == 'ReasonforProcessingOthers'):
            # Create a data validation object
            dv = DataValidation(type="list", formula1='INDIRECT(SUBSTITUTE(P1," ","_")&"_CATEGORY")', allow_blank=True,
                                error='Your entry is not valid',errorTitle='Invalid Entry', 
                                prompt='Please select from the list', promptTitle='Select Option')            
            # Add the data validation to the worksheet
            ws.add_data_validation(dv)
            # Apply the data validation to the column
            for row in ws[rows['Position']]:
                dv.add(row)                
        # Formating the main sheet
        my_fill = PatternFill(start_color="C0C0C0", fill_type="solid")
        for rows in ws.iter_rows(min_row=1, max_row=1, min_col=None):
            for cell in rows:
                cell.font = Font(bold=True )
                cell.fill = my_fill                
        ws.freeze_panes = ws['B2']
        ws.protection.sheet = True
        ws.protection.password = password
        ws.protection.enable()
        for col in list(mapping_df['Position']):
            for cell in ws[col]:
                cell.protection = Protection(locked=False)                
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width        
        hidden_ws.sheet_state = 'hidden'       
        # Save the workbook
        wb.save('C:\\Users\\2014201\\Downloads\\' + LOB+'.xlsx')        
    return


def set_wb_password(filePath, password):
    xl_file = EnsureDispatch("Excel.Application")
    wb = xl_file.Workbooks.Open(filePath)
    xl_file.DisplayAlerts = False
    xl_file.Visible = False
    wb.SaveAs(filePath, Password = password)
    wb.Close()
    xl_file.Quit()
    
        
        

path = 'C:\\Users\\2014201\Downloads\\'         
AttributesFileName = 'Attributes.xlsx'
MappingFileName = 'Attribute_Mapping.xlsx'

AtrributesFilePath = path + AttributesFileName
MappingFilePath = path + MappingFileName
password = 'WFM_GBS'
                               
attribute_df = pd.read_excel(AtrributesFilePath)

DropDownMapping_df = pd.read_excel(MappingFilePath, sheet_name='DropDownMapping')
ColumnMapping_df = pd.read_excel(MappingFilePath, sheet_name='ColumnMapping')

attribute_df['LOB'] = attribute_df['LOB'].str.upper()

listOfLOBs = list(set(attribute_df['LOB'].str.upper().dropna()))

for LOB in listOfLOBs:
    f_attribute_df = attribute_df.loc[attribute_df['LOB'].str.contains(LOB, case=False,na=False)]
    add_data_validation_to_column(LOB, f_attribute_df, ColumnMapping_df,DropDownMapping_df, password)
#     set_wb_password(path+LOB+'.xlsx', password)



from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.protection import Protection

def add_data_validation_to_column(LOB, attribute_df, mapping_df, DropDownMapping_df, password):
    """
    This function adds data validation to a column in an Excel worksheet.
    It takes as input the following parameters:
    - LOB: Line of Business
    - attribute_df: DataFrame containing attributes
    - mapping_df: DataFrame containing mapping information
    - DropDownMapping_df: DataFrame containing dropdown mapping information
    - password: Password for sheet protection
    """
    
    # Create a workbook and add a worksheet to it
    wb = Workbook()
    ws = wb.create_sheet(LOB, 0)
    
    # Create a hidden worksheet for storing mapping information
    hidden_ws = wb['Sheet']
    hidden_ws.title = 'MappingSheet'
    
    # Add rows from attribute_df and DropDownMapping_df to the worksheets
    for r in dataframe_to_rows(attribute_df, index=False, header=True):
        ws.append(r)
        
    for r in dataframe_to_rows(DropDownMapping_df, index=False, header=True):
        hidden_ws.append(r)
        
    # Define a dictionary with named range titles as keys and their corresponding cell ranges as values   
    named_ranges = {
        'PROCESSING_OTHERS': '$B$2:$B$6',
        'PROCESSING': '$B$7:$B$10',
        'NON_PROCESSING': '$B$11:$B$20',
        'PROCESSING_OTHERS_CATEGORY': '$C$2:$C$6'
    }
    
    # Create named ranges in the hidden worksheet for data validation
    for title, cell_range in named_ranges.items():
        wb.create_named_range(title, hidden_ws, cell_range)
    
    # Filter the mapping DataFrame to get rows where DropDownOptions is not null
    f_mapping_df = mapping_df[mapping_df['DropDownOptions'].notnull()]
    
    # Define a dictionary with DropDownOptions and Columns as keys and their corresponding formulas as values
    data_validations = {
        ('Named Range', 'Role'): 'INDIRECT(SUBSTITUTE(P1," ","_"))',
        ('Named Range', 'ReasonforProcessingOthers'): 'INDIRECT(SUBSTITUTE(P1," ","_")&"_CATEGORY")'
    }
    
    # Iterate over the filtered DataFrame and add data validation to the columns based on the DropDownOptions value
    for index, row in f_mapping_df.iterrows():
        if row['DropDownOptions'] != 'Named Range':
            formula = '\"' + str(row['DropDownOptions']) + '\"'
        else:
            formula = data_validations.get((row['DropDownOptions'], row['Columns']), None)
            
        if formula is not None:
            # Create a data validation object with list type and formula1 as formula
            dv = DataValidation(type="list", formula1=formula, error='Your entry is not valid',
                                errorTitle='Invalid Entry', prompt='Please select from the list', 
                                promptTitle='Select Option')
            
            # Add the data validation to the worksheet and apply it to the column
            ws.add_data_validation(dv)
            for cell in ws[row['Position']]:
                dv.add(cell)

                
        # Define the font and fill styles
        bold_font = Font(bold=True)
        my_fill = PatternFill(start_color="C0C0C0", fill_type="solid")
        
        # Apply the styles to the cells in the header row
        for cell in ws[1]:
            cell.font = bold_font
            cell.fill = my_fill
                
        ws.freeze_panes = ws['B2']
        
        # Protect the worksheet with a password and enable protection
        ws.protection.sheet = True
        ws.protection.password = password
        ws.protection.enable()
        
        # Set protection for cells in the columns specified in mapping_df
        for col in list(mapping_df['Position']):
            for cell in ws[col]:
                cell.protection = Protection(locked=False)
                
    # Iterate over each column in the worksheet
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        # Try to find the maximum length of cell value in the column
        try:
            max_length = max(len(str(cell.value)) for cell in column)
        except:
            pass
        
        # Adjust the width of the column based on the maximum length of cell value
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
            
    return wb

def set_wb_password(file_path, password):
    try:
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False

        wb = excel.Workbooks.Open(file_path)
        wb.SaveAs(file_path, Password=password)
        
    except Exception as e:
        print(f"An error occurred: {e}")
        
    finally:
        wb.Close()
        excel.Quit()

from openpyxl.security import encrypt_archive

def password_protect_excel_file(excel_file, password):
    # Load the excel file
    wb = openpyxl.load_workbook(excel_file)

    # Apply the password to the workbook
    wb.security = encrypt_archive(password)

    # Save the password protected excel file
    wb.save(excel_file)

# Usage example
password_protect_excel_file('my_excel_file.xlsx', 'my_password')

# Define the file path and names
path = 'C:\\Users\\2014201\\Downloads\\'
attributes_file_name = 'Attributes.xlsx'
mapping_file_name = 'Attribute_Mapping.xlsx'

# Use os.path.join to create the file paths
attributes_file_path = os.path.join(path, attributes_file_name)
mapping_file_path = os.path.join(path, mapping_file_name)

password = 'WFM_GBS'

# Read the Excel files into DataFrames
attribute_df = pd.read_excel(attributes_file_path)
drop_down_mapping_df = pd.read_excel(mapping_file_path, sheet_name='DropDownMapping')
column_mapping_df = pd.read_excel(mapping_file_path, sheet_name='ColumnMapping')

# Convert the 'LOB' column to uppercase
attribute_df['LOB'] = attribute_df['LOB'].str.upper()

# Get the unique LOBs
list_of_lobs = attribute_df['LOB'].dropna().unique()

# Iterate over the LOBs and add data validation to each one
for lob in list_of_lobs:
    f_attribute_df = attribute_df[attribute_df['LOB'].str.contains(lob, case=False, na=False)]
    add_data_validation_to_column(lob, f_attribute_df, column_mapping_df, drop_down_mapping_df, password)

from openpyxl.utils import get_column_letter

def get_excel_column_letter_from_value(file_name, target_value):
    workbook = load_workbook(filename=file_name)
    sheet = workbook.active

    for cell in sheet[1]:  # Iterate over the cells in the first row
        if cell.value == target_value:
            return get_column_letter(cell.column)  # Return the column letter if the cell value matches the target value

    return None  # Return None if no match is found

# Test the function
print(get_excel_column_letter_from_value('your_file.xlsx', 'your_target_value'))
