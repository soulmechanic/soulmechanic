from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

def add_data_validation_to_column(file_path, sheet_name, column_letter):
    # Load the workbook and select the sheet
    wb = load_workbook(file_path)
    ws = wb[sheet_name]

    # Create a data validation object
    dv = DataValidation(type="whole", operator="between", formula1=0, formula2=100)

    # Apply the data validation to the column
    for row in ws[column_letter]:
        dv.add(row)

    # Add the data validation to the worksheet
    ws.add_data_validation(dv)

    # Save the workbook
    wb.save(file_path)

add_data_validation_to_column('path_to_your_file.xlsx', 'Sheet1', 'A')
