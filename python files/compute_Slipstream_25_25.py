
# coding: utf-8

# In[1]:


# -*- coding: utf-8 -*-
import dataiku
from dataiku import pandasutils as pdu
import pyarrow.parquet as pq
import pandas as pd, numpy as np
import json
import io
from io import BytesIO
import re
from zipfile import ZipFile
import zipfile
from pandas.io.json import json_normalize
import xlrd
from openpyxl import load_workbook
from string import ascii_lowercase

from office365.runtime.auth.client_credential import ClientCredential
from office365.runtime.auth.user_credential import UserCredential
from office365.runtime.http.request_options import RequestOptions
from office365.sharepoint.client_context import ClientContext
from office365.sharepoint.listitems.caml.caml_query import CamlQuery 
from office365.sharepoint.listitems.listitem import ListItem
from office365.sharepoint.fields.field_creation_information import FieldCreationInformation
from office365.sharepoint.fields.field_multi_user_value import FieldMultiUserValue
from office365.sharepoint.fields.field_type import FieldType
from office365.sharepoint.fields.field_user_value import FieldUserValue
from office365.sharepoint.lists.list_creation_information import ListCreationInformation
from office365.sharepoint.lists.list_template_type import ListTemplateType


# In[2]:


# Read recipe inputs
slipstream_S3_Bucket = dataiku.Folder("ta9nWjmk")
slipstream_S3_Bucket_info = slipstream_S3_Bucket.get_info()


# ### Read Excel from S3 Bucket 

# In[3]:


def extract_excel_data(excel_folder, input_excel, worksheet):
    try:
        if input_excel[0].endswith('.xlsx'):

            #Load a Template
            with excel_folder.get_download_stream(input_excel) as f:
               templateFile = f.read()

            load_WB = load_workbook(BytesIO(templateFile))
            WorkSheet = load_WB[worksheet]
            data = WorkSheet.values

            # Get the first line in file as a header line
            columns = next(data)[0:]

            # Convert to DataFrame
            OutDF = pd.DataFrame(WorkSheet.values, columns=columns)
            OutDF = OutDF.iloc[1: , :].reset_index(drop=True)

        return OutDF

    except Exception as e:
            print(input_excel[0]+': is not excel file')


# In[4]:


# Search for that particular excel file in S3 folder
S3_excel_folder = slipstream_S3_Bucket
excel_file = 'Table_25x25.xlsx'
worksheet = '25_25'

input_excel = [x for x in S3_excel_folder.list_paths_in_partition() if re.search(excel_file,x)]

# calling the extract_excel_data function to read slipstream.xlsx
Excel_From_S3_df = extract_excel_data(S3_excel_folder, input_excel, worksheet)
# Excel_From_S3_df


# ### Read SharePoint List and Convert to Dataframe

# In[5]:


def auth_sharepoint():
    ## supplying the credential for accessing the SharePoint site.
    ## Credentials are save in Global Variables for this project.
    site_url = "https://pfizer.sharepoint.com/" #sharepoint tenant url
    OLS = 'sites/PortfolioDev/scratch'#name of the site

    xUsername=dataiku.get_custom_variables()["Username"]
    xPassword=dataiku.get_custom_variables()["Password"]

    ctx = ClientContext(site_url+OLS).with_credentials(UserCredential(xUsername, xPassword))
    return ctx


# In[6]:


ctx = auth_sharepoint()

## read sharepoint list and convert to dataframe
OLN = 'Slipstream_25by25'
sharepoint_list = ctx.web.lists.get_by_title(OLN)
all_items = sharepoint_list.items.get().execute_query()

#Read from the SharePoint List
final_list = []
for item in all_items:
    final_list.append({'CandidateCode':item.properties['CandidateCode'],
                    'ValuationPhase': item.properties['Title'],
                   'DiseaseAreaLong':item.properties['DiseaseAreaLong'],
                   'FundingStatus':item.properties['FundingStatus'],
                   'ParentChild':item.properties['ParentChild'],
                   'ProductName':item.properties['ProductName'],
                   'TherapeuticArea':item.properties['TherapeuticArea'],
                   'TwentyFiveby25Program':item.properties['TwentyFiveby25Program'],
                   'Year':item.properties['Year']})
    my_json = json.dumps(final_list)
SharePoint_List_df = pd.json_normalize(json.loads(my_json))
SharePoint_List_df['Year'] = SharePoint_List_df['Year'].map(str)
# SharePoint_List_df


# In[7]:


SharePoint_List_df


# ###  Write, Update and Clear to a sharepoint list 

# #### Clear data in sharepoint list

# In[8]:


# function to clear existing sharepoint list
def clear_SPList(SPList):
    # clearing the sharepoint list data based on condition
    ctx = auth_sharepoint()
    target_list = ctx.web.lists.get_by_title(SPList)
    items = target_list.items.get().execute_query()
    for item in items:  # type: ListItem
        item.delete_object()
    ctx.execute_batch()
    print("Items deleted count: {0}".format(len(items)))
    
    
clear_SPList("Slipstream_25by25_V2")


# #### Upload data from Dataframe to Sharepoint list

# In[9]:


## adding data to sharepoint list from a dataframe
ctx = auth_sharepoint()



target_list = ctx.web.lists.get_by_title("Slipstream_25by25_V2")

for index, row in SharePoint_List_df.iterrows():

    target_list.add_item({'Title': row['ValuationPhase'], 'CandidateCode': row['CandidateCode'], 
                         'DiseaseAreaLong': row['DiseaseAreaLong'], 'FundingStatus': row['FundingStatus'],
                        'ParentChild': row['ParentChild'], 'ProductName': row['ProductName'],
                        'TherapeuticArea': row['TherapeuticArea'], 'TwentyFiveby25Program': row['TwentyFiveby25Program'],
                        'Year': row['Year']})
ctx.execute_batch()


# #### update existing sharepoint list data on a particular column based on condition

# In[10]:


# create a scenrio where tranformation or changes to data is done and this is updated to Sharepoint list
SharePoint_List_df.loc[(SharePoint_List_df.ValuationPhase == 'Approved'),'TwentyFiveby25Program']='No'
F_SharePoint_List_df = SharePoint_List_df[SharePoint_List_df['ValuationPhase']=='Approved']
F_SharePoint_List_df

target_list = ctx.web.lists.get_by_title("Slipstream_25by25_V2")
# updating the sharepoint list column TwentyFiveby25Program with only those with 'ValuationPhase' column has 'approved' as values
qry_text ='''<Where>
                <Eq> 
                   <FieldRef Name='Title'/>
                   <Value Type='Text'>Approved</Value>
                </Eq>          
              </Where>'''
caml_query =CamlQuery.parse(qry_text)

items = target_list.get_items(caml_query) 
items = target_list.items
ctx.load(items).execute_query()

for row, item in zip(F_SharePoint_List_df.itertuples(index=True, name='Pandas'), items):
    item.set_property("TwentyFiveby25Program", row.TwentyFiveby25Program).update()
ctx.execute_batch()


# #### Creating SharePoint list from scratch - this is work in progress

# In[38]:


def ensure_list(web, list_properties):
    lists = web.lists.filter("Title eq '{0}'".format(list_properties.Title)).get().execute_query()
    if len(lists) == 1:
        return lists[0]
    else:
        return web.lists.add(list_properties).execute_query()
    
    
def create_list(list_title,cols):
    ctx = auth_sharepoint()

    target_list = ensure_list(ctx.web,
                                ListCreationInformation(list_title,
                                                        None,
                                                        ListTemplateType.GenericList)
                                )

    field_info = FieldCreationInformation("ValuationPhase", field_type_kind=FieldType.Text)
    user_field = target_list.fields.add(field_info).execute_query()
    return target_list


# In[39]:


create_list("Slipstream_25by25_temp",SharePoint_List_df.columns)


# ### Dwonloading Excel and Read from SharePoint Document Library

# In[7]:


def excel_extract(file,worksheet):
    xlfile = load_workbook(BytesIO(file.read()))
    WorkSheet = xlfile[worksheet]
    data = WorkSheet.values
    # Get the first line in file as a header line
    columns = next(data)[0:]
    # Convert to DataFrame
    OutDF = pd.DataFrame(WorkSheet.values, columns=columns)
    OutDF = OutDF.iloc[1: , :].reset_index(drop=True)

    return OutDF


# In[8]:


SP_SITE_URL ='https://pfizer.sharepoint.com/sites/PortfolioDev/scratch/'
SP_DOC_LIBRARY ='Slipstream_lib'

# Create a ClientContext object and use the user’s credentials for authentication 
ctx =ClientContext(SP_SITE_URL).with_user_credentials(xUsername, xPassword) 


# In[9]:


# Build a CAML query to find all list items that associate the approved articles 
qry_text ='''<Where>
                <Eq> 
                   <FieldRef Name='Title'/>
                   <Value Type='Text'>Approved</Value>
                </Eq>          
              </Where>'''
caml_query =CamlQuery.parse(qry_text)
caml_query.FolderServerRelativeUrl = SP_DOC_LIBRARY 

# Retrieve list items based on the CAML query 
oList = ctx.web.lists.get_by_title(SP_DOC_LIBRARY)
items = oList.get_items(caml_query) 
ctx.execute_query()

# Loop through all list items
for item in items:
    # Query the SharePoint document library to find the associated file of a list item 
    file = item.file
    ctx.load(file)
    ctx.execute_query()
    Excel_From_SP_DL_DF = excel_extract(file,'25_25')
Excel_From_SP_DL_DF


# ### Uploading file to SharePoint Library

# In[10]:


SP_DOC_LIBRARY_FOLDER ='Upload_folder'

# Access the SharePoint folder or create the folder if it doesn’t exist 
sp_folder_path ="/{0}/{1}".format(SP_DOC_LIBRARY,SP_DOC_LIBRARY_FOLDER)
sp_folder  =  ctx.web.ensure_folder_path(sp_folder_path).execute_query()

#########################################################################################################
# converting Dataframe to json 
_json = json.dumps([row.dropna().to_dict() for index,row in Excel_From_SP_DL_DF.iterrows()],
                   indent=1,ensure_ascii=False,default=str)

# tidy json and add metadata wrapper
_json = _json.replace('\n','\r\n')

_json = '{"PDA":'+ _json +'}'

#########################################################################################################

# Upload the file to the folder in the SharePoint library 
sp_file = sp_folder.upload_file('TwentyFiveby25.txt', _json.encode('utf-8')) 
ctx.execute_query()


# ### Writeout to Output

# In[10]:


# Compute recipe outputs
# TODO: Write here your actual code that computes the outputs
# NB: DSS supports several kinds of APIs for reading and writing data. Please see doc.

slipstream_25_25_df = SharePoint_List_df # Compute a Pandas dataframe to write into Slipstream_25_25


# Write recipe outputs
slipstream_25_25 = dataiku.Dataset("Slipstream_25_25")
slipstream_25_25.write_with_schema(slipstream_25_25_df)


# #### CAML Query Samples

# 1. Null:
# <Where><IsNull><FieldRef Name="CustomField" /></IsNull></Where>
# 2. Not Null:
# <Where><IsNotNull><FieldRef Name="CustomField" /></IsNotNull></Where>
# 3. Equal:
# <Where><Eq><FieldRef Name="CustomField" /><Value Type="Text">MatchValue</Value></Eq></Where>
# 4. Not Equal：
# <Where><Neq><FieldRef Name="CustomField" /><Value Type="Text">MatchValue</Value></Neq></Where>
# 5. Greater Than：
# <Where><Gt><FieldRef Name="CustomField" /><Value Type="Text">1</Value></Gt></Where>
# 6. Greater Than And Equal：
# <Where><Geq><FieldRef Name="CustomField" /><Value Type="Text">1</Value></Geq></Where>
# 7. Lower Than:
# <Where><Lt><FieldRef Name="CustomField" /><Value Type="Text">1</Value></Lt></Where>
# 8. Lower Than And Equal:
# <Where><Leq><FieldRef Name="CustomField" /><Value Type="Text">1</Value></Leq></Where>
# 9 Begin With:
# <Where><BeginsWith><FieldRef Name="CustomField" /><Value Type="Text">StartString</Value></BeginsWith></Where>
# 10: Contains:
# <Where><Contains><FieldRef Name="CustomField" /><Value Type="Text">ContainString</Value></Contains></Where>

# #### OLDER CODE

# In[ ]:


# qry_text ='''<Where>
#                 <IsNotNull> 
#                    <FieldRef Name='CandidateCode'/>
#                 </IsNotNull>
#               </Where>'''
# caml_query =CamlQuery.parse(qry_text)

# items = list_tasks.get_items(caml_query) 
# # ctx.execute_query()
# for item in items:
#     ctx.load(item) 
#     item.delete_object()
# ctx.execute_query() 


# items_to_add = [target_list.add_item(Add_Items) for index,row in df.iterrows()]

